from tkinter import*
import tkinter.messagebox as messagebox
from tkinter.messagebox import askyesno
from time import sleep
import os
from datetime import datetime
import openpyxl 
try:
    wb_obj = openpyxl.load_workbook("SystemUsageData.xlsx")
except FileNotFoundError:
    wb_obj = openpyxl.Workbook()

sheet_obj = wb_obj.active
l_row = sheet_obj.max_row
column = sheet_obj.max_column

root = Tk()

# user define function
def shutdown():
	return os.system("shutdown /s /t 1")

def restart():
	return os.system("shutdown /r /t 1")

def logout():
	return os.system("shutdown -l")

def submit():
    now = datetime.now()
    
    sheet_obj.cell(l_row+1,1).value=l_row
    sheet_obj.cell(l_row+1,2).value=now
    sheet_obj.cell(l_row+1,3).value=ename.get()
    sheet_obj.cell(l_row+1,4).value=eemail.get()
    sheet_obj.cell(l_row+1,5).value="male" if egender.get() == 1 else "female"
    sheet_obj.cell(l_row+1,6).value=eage.get()
    sheet_obj.cell(l_row+1,7).value=eTime.get()
    wb_obj.save("SystemUsageData.xlsx")
    TIMELIMTI = eTime.get()
    root.destroy()
    messagebox.showinfo("Thank You",  "Now Start Working!!")
    sleep(3600*eval(TIMELIMTI))
    shutdown()
    
def confirmExit():
    answer = askyesno(title='confirmation',
                    message='Are you sure that you want to quit?')
    if answer:
        shutdown()
    
    
def validate():
    if not ename.get():
        messagebox.showinfo("Warning!!.",  "Enter Your Full Name!")
    elif not eemail.get():
        messagebox.showinfo("Warning!!.",  "Enter Your Email!")
    elif not egender.get():
        messagebox.showinfo("Warning!!.",  "Enter Your Gender!")
    elif not eage.get():
        messagebox.showinfo("Warning!!.",  "Enter Your Age!")
    elif not eTime.get():
        messagebox.showinfo("Warning!!.",  "Enter Time Limit!")
    else:
        submit()
        
        
root.attributes("-fullscreen", True)
root.title("Registration Form")
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

heading = Label(root, text="Registration form",width=20,font=("bold", 20))
heading.place(x=screen_width/2.5,y=53)
    

lname = Label(root, text="FullName",width=20,font=("bold", 10))
lname.place(x=screen_width/2.5,y=130)

ename = Entry(root) #Name
ename.place(x=screen_width/2.5+130,y=130)

lemail = Label(root, text="Email",width=20,font=("bold", 10))
lemail.place(x=screen_width/2.5,y=180)

eemail = Entry(root) #Email
eemail.place(x=screen_width/2.5+130,y=180)

lgender = Label(root, text="Gender",width=20,font=("bold", 10))
lgender.place(x=screen_width/2.5,y=230)
egender = IntVar()
Radiobutton(root, text="Male",padx = 5, variable=egender, value=1).place(x=screen_width/2.5+130,y=230)
Radiobutton(root, text="Female",padx = 20, variable=egender, value=2).place(x=screen_width/2.5+185,y=230)

lage = Label(root, text="Age:",width=20,font=("bold", 10))
lage.place(x=screen_width/2.5,y=280)
eage = Entry(root) #Age
eage.place(x=screen_width/2.5+130,y=280)

lTime = Label(root,text="Time Limit(hr):",font=("bold", 10))
lTime.place(x=screen_width/2.5,y=330)
eTime = Entry(root) #TimeLimit
eTime.place(x=screen_width/2.5+130,y=330)


Button(root, text='Submit',width=20,bg='brown',fg='white',command=validate).place(x=screen_width/2.25-70,y=430)
# it is use for display the registration form on the window
Button(root, text='Exit',width=20,bg='brown',fg='white',command=confirmExit).place(x=screen_width/2.25+70,y=430)
root.mainloop()
